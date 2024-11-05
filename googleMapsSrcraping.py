import os
import time
from bs4 import BeautifulSoup
from fpdf import FPDF
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook

# Função para garantir que o texto seja codificado corretamente
def safe_text(text):
    return text.encode('latin-1', 'replace').decode('latin-1')

# Configurar o Selenium WebDriver
options = Options()
# Remover a linha abaixo para não executar em modo headless
# options.add_argument('--headless')  # Executa o Chrome em modo headless (sem abrir o navegador)
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# URL do site com as avaliações
url = 'https://www.google.com/maps/place/Hospital+de+Cl%C3%ADnicas/@-25.4241557,-49.2645227,17z/data=!4m8!3m7!1s0x94dce43e77959da7:0x725e3a407b4e2f45!8m2!3d-25.4241606!4d-49.2619478!9m1!1b1!16s%2Fg%2F1tdzbklv?entry=ttu&g_ep=EgoyMDI0MTAyMy4wIKXMDSoASAFQAw%3D%3D'
# Acessar a página
driver.get(url)
time.sleep(5)  # Esperar o carregamento inicial da página

# Definir o limite de avaliações
limite_avaliacoes = 350  # Coloque aqui o número máximo de avaliações que você deseja capturar

# Lista para armazenar todas as avaliações
avaliacoes = []

# Loop para continuar o scroll até que o número máximo de avaliações seja atingido
while len(avaliacoes) < limite_avaliacoes:
    # Scroll para carregar mais avaliações
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(5)  # Aguardar o carregamento após o scroll

    # Contar o número de avaliações visíveis
    novas_avaliacoes = driver.find_elements(By.CLASS_NAME, 'jftiEf')
    
    # Adicionar novas avaliações à lista
    for avaliacao in novas_avaliacoes:
        if avaliacao not in avaliacoes:
            avaliacoes.append(avaliacao)
    
    # Verificar se novas avaliações foram carregadas
    if len(novas_avaliacoes) == 0:
        break  # Parar o loop se não houver mais avaliações sendo carregadas

# Criar um objeto PDF
pdf = FPDF()
pdf.set_auto_page_break(auto=True, margin=15)
pdf.add_page()
pdf.set_font("Arial", size=12)

# Criar um objeto Workbook para o Excel
wb = Workbook()
ws = wb.active
ws.title = "Avaliações"

# Adicionar cabeçalhos ao Excel
ws.append(["Nome", "Nota", "Data", "Comentário", "Resposta da Empresa"])

# Percorrer as avaliações e extrair as informações (limitando ao número desejado)
for avaliacao in avaliacoes[:limite_avaliacoes]:  # Limitar ao número máximo de avaliações
    try:
        # Clique no botão "Mais" se ele estiver presente
        try:
            mais_button = avaliacao.find_element(By.CLASS_NAME, 'w8nwRe.kyuRq')  # Classe do botão "Mais"
            mais_button.click()
            time.sleep(1)  # Esperar um momento para o texto expandido ser carregado
        except:
            pass  # Se não houver botão "Mais", siga em frente

        # Nome do paciente
        nome = avaliacao.find_element(By.CLASS_NAME, 'd4r55').text.strip()
    except:
        nome = 'Nome não encontrado'

    try:
        # Nota do paciente
        nota = avaliacao.find_element(By.CLASS_NAME, 'kvMYJc').get_attribute('aria-label')
    except:
        nota = 'Nota não encontrada'

    try:
        # Comentário do paciente (depois do clique no "Mais", se houver)
        comentario = avaliacao.find_element(By.CLASS_NAME, 'MyEned').text.strip()
    except:
        comentario = 'Comentário não encontrado'

    try:
        # Data do comentário
        data_comentario = avaliacao.find_element(By.CLASS_NAME, 'rsqaWe').text.strip()
    except:
        data_comentario = 'Data não encontrada'

    try:
        # Resposta da empresa
        resposta_empresa = avaliacao.find_element(By.XPATH, './/div[contains(@class, "wiI7pd")]').text.strip()
    except:
        resposta_empresa = 'Resposta não encontrada'

    # Escrever no PDF (convertendo o texto para o encoding correto)
    pdf.cell(200, 10, txt=safe_text(f"Nome: {nome}"), ln=True, align='L')
    pdf.cell(200, 10, txt=safe_text(f"Nota: {nota}"), ln=True, align='L')
    pdf.cell(200, 10, txt=safe_text(f"Data: {data_comentario}"), ln=True, align='L')
    pdf.multi_cell(200, 10, txt=safe_text(f"Comentário: {comentario}"), align='L')
    pdf.multi_cell(200, 10, txt=safe_text(f"Resposta da empresa: {resposta_empresa}"), align='L')
    pdf.cell(200, 10, txt=safe_text('-' * 40), ln=True, align='L')

    # Adicionar os dados ao Excel
    ws.append([nome, nota, data_comentario, comentario, resposta_empresa])

# Caminho para salvar o arquivo na pasta Documentos
documents_path = os.path.join(os.path.expanduser("~"), "Documents")

# Verificar se o diretório existe, se não, criar
if not os.path.exists(documents_path):
    os.makedirs(documents_path)

# Salvar o PDF na pasta Documentos
pdf_path = os.path.join(documents_path, "avaliacoes.pdf")
pdf.output(pdf_path)
print(f"PDF salvo com sucesso em: {pdf_path}")

# Salvar o Excel na pasta Documentos
excel_path = os.path.join(documents_path, "avaliacoes.xlsx")
wb.save(excel_path)
print(f"Excel salvo com sucesso em: {excel_path}")

# Fechar o navegador
driver.quit()